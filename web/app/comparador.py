"""
Core de comparação de planilhas do RH.
Portado e adaptado de comparador_rh_v2.py para uso web.
"""

import re
import shutil
import json
import unicodedata
from datetime import datetime
from pathlib import Path
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Configuração de colunas ────────────────────────────────────────────────
CONFIG = {
    "col_matricula":    "MATRÍCULA",
    "col_nome":         "NOME",
    "col_cargo":        "CARGO",
    "col_departamento": "DEPARTAMENTO",
    "col_gestor":       "GESTOR",
    "col_admissao":     "ADMISSÃO",
}

CAMPOS_MONITORADOS = ["CARGO", "DEPARTAMENTO", "GESTOR"]

# Mapeamento para chaves limpas no JSON (sem acentos)
CAMPO_MAP = {
    "TIPO":          "tipo",
    "MATRÍCULA":     "matricula",
    "NOME":          "nome",
    "CARGO":         "cargo",
    "DEPARTAMENTO":  "departamento",
    "GESTOR":        "gestor",
    "ADMISSÃO":      "admissao",
}

# ── Estilos Excel ──────────────────────────────────────────────────────────
FONT_TITLE   = Font(name="Arial", size=13, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_HEADER  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_BODY    = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", size=10, bold=True)

FILL_TITLE  = PatternFill("solid", start_color="263238")
FILL_ADD    = PatternFill("solid", start_color="2E7D32")
FILL_REM    = PatternFill("solid", start_color="C62828")
FILL_ALT    = PatternFill("solid", start_color="EF6C00")
FILL_HEADER = PatternFill("solid", start_color="455A64")
FILL_ZEBRA  = PatternFill("solid", start_color="ECEFF1")
FILL_DIFF   = PatternFill("solid", start_color="FFF59D")

THIN   = Side(border_style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Carga e normalização ───────────────────────────────────────────────────

def carregar_planilha(caminho: Path) -> pd.DataFrame:
    """Carrega, normaliza e indexa uma planilha do RH."""
    df = pd.read_excel(caminho, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]

    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"NAN": "", "NONE": "", "NAT": ""})

    # Descarta linhas em branco (sem matrícula E sem nome)
    col_m = CONFIG["col_matricula"]
    col_n = CONFIG["col_nome"]
    df = df[~((df[col_m] == "") & (df[col_n] == ""))].copy()

    def gerar_chave(row):
        matr_raw = str(row.get(col_m, "")).strip()
        nome = str(row.get(col_n, "")).strip().upper()
        matr = matr_raw.upper()
        if matr.endswith(".0") and matr[:-2].replace("-", "").isdigit():
            matr = matr[:-2]
        if matr in ("PJ", "", "NAN", "NONE"):
            return f"PJ::{nome}" if nome else f"VAZIA::{id(row)}"
        return f"CLT::{matr}"

    df["_CHAVE"] = df.apply(gerar_chave, axis=1)

    def limpar_matr(v):
        v = str(v).strip()
        if v.endswith(".0") and v[:-2].replace("-", "").isdigit():
            return v[:-2]
        return v

    df[col_m] = df[col_m].apply(limpar_matr)

    dup = df[df["_CHAVE"].duplicated(keep=False)]
    if not dup.empty:
        df = df.drop_duplicates(subset="_CHAVE", keep="last")

    return df.set_index("_CHAVE")


def comparar(df_antigo: pd.DataFrame, df_novo: pd.DataFrame):
    """Retorna (adicoes, remocoes, alteracoes) como DataFrames."""
    k_ant, k_nov = set(df_antigo.index), set(df_novo.index)

    adicoes  = df_novo.loc[list(k_nov - k_ant)].copy()
    remocoes = df_antigo.loc[list(k_ant - k_nov)].copy()

    linhas_alt = []
    for chave in k_ant & k_nov:
        antigo = df_antigo.loc[chave]
        novo   = df_novo.loc[chave]
        diffs  = {}
        for campo in CAMPOS_MONITORADOS:
            v_ant = str(antigo.get(campo, "")).strip()
            v_nov = str(novo.get(campo, "")).strip()
            if v_ant != v_nov:
                diffs[campo] = (v_ant, v_nov)
        if diffs:
            linha = {
                "TIPO":     "PJ" if chave.startswith("PJ::") else "CLT",
                "MATRÍCULA": novo.get(CONFIG["col_matricula"], ""),
                "NOME":     novo.get(CONFIG["col_nome"], ""),
            }
            for campo in CAMPOS_MONITORADOS:
                if campo in diffs:
                    linha[f"{campo}_ANTES"]  = diffs[campo][0]
                    linha[f"{campo}_DEPOIS"] = diffs[campo][1]
                    linha[f"{campo}_MUDOU"]  = True
                else:
                    v = novo.get(campo, "")
                    linha[f"{campo}_ANTES"]  = v
                    linha[f"{campo}_DEPOIS"] = v
                    linha[f"{campo}_MUDOU"]  = False
            linhas_alt.append(linha)

    alteracoes = pd.DataFrame(linhas_alt)

    for df in (adicoes, remocoes):
        if not df.empty:
            df.insert(0, "TIPO",
                      ["PJ" if k.startswith("PJ::") else "CLT" for k in df.index])

    return adicoes, remocoes, alteracoes


def _df_to_list(df: pd.DataFrame, campos: list) -> list:
    """Converte DataFrame para lista de dicts com chaves limpas (sem acentos)."""
    result = []
    for _, row in df.iterrows():
        item = {}
        for campo in campos:
            valor = row.get(campo, "")
            clean = CAMPO_MAP.get(campo, campo.lower())
            item[clean] = str(valor) if pd.notna(valor) else ""
        result.append(item)
    return result


def extrair_data(arquivo: Path) -> str:
    # Usa apenas "." ou "-" como separadores (não "_") para evitar capturar
    # o timestamp prefixado no nome do arquivo (ex: 20260527_123626_efetivo-26.05.xlsx
    # teria "27_12" como falso-positivo se "_" fosse separador aceito).
    m = re.search(r"(\d{2})[.\-](\d{2})(?:[.\-](\d{2,4}))?", arquivo.stem)
    if not m:
        return datetime.now().strftime("%d-%m-%Y")
    dd, mm, yy = m.group(1), m.group(2), m.group(3)
    return f"{dd}-{mm}-{yy}" if yy else f"{dd}-{mm}"


# ── Processamento principal ────────────────────────────────────────────────

def processar_planilha(
    arquivo_novo: Path,
    pasta_snapshots: Path,
    pasta_relatorios: Path,
    usuario: str = "sistema",
) -> dict:
    """
    Processa uma nova planilha do RH:
    1. Carrega e normaliza o arquivo
    2. Compara com o snapshot mais recente (se existir)
    3. Salva resultado em JSON
    4. Salva snapshot do arquivo para próxima comparação

    Retorna dict com todos os dados da comparação.
    """
    pasta_snapshots.mkdir(parents=True, exist_ok=True)
    pasta_relatorios.mkdir(parents=True, exist_ok=True)

    df_novo = carregar_planilha(arquivo_novo)
    n_pj  = sum(1 for k in df_novo.index if k.startswith("PJ::"))
    n_clt = sum(1 for k in df_novo.index if k.startswith("CLT::"))

    snapshots = sorted(pasta_snapshots.glob("snapshot_*.xlsx"))
    snapshot_anterior = snapshots[-1] if snapshots else None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    data_str  = extrair_data(arquivo_novo)

    resultado: dict = {
        "id":        timestamp,
        "data":      data_str,
        "arquivo":   arquivo_novo.name,
        "timestamp": datetime.now().isoformat(),
        "usuario":   usuario,
        "total":     len(df_novo),
        "total_clt": n_clt,
        "total_pj":  n_pj,
        "is_baseline": snapshot_anterior is None,
        "n_adicoes":    0,
        "n_remocoes":   0,
        "n_alteracoes": 0,
        "adicoes":    [],
        "remocoes":   [],
        "alteracoes": [],
    }

    if snapshot_anterior is not None:
        df_antigo = carregar_planilha(snapshot_anterior)
        adicoes, remocoes, alteracoes = comparar(df_antigo, df_novo)

        campos_adic = ["TIPO", "MATRÍCULA", "NOME", "CARGO", "DEPARTAMENTO", "GESTOR", "ADMISSÃO"]
        campos_rem  = ["TIPO", "MATRÍCULA", "NOME", "CARGO", "DEPARTAMENTO", "GESTOR"]

        resultado["n_adicoes"]  = len(adicoes)
        resultado["n_remocoes"] = len(remocoes)
        resultado["n_alteracoes"] = len(alteracoes)

        resultado["adicoes"]  = _df_to_list(adicoes,  campos_adic) if not adicoes.empty else []
        resultado["remocoes"] = _df_to_list(remocoes, campos_rem)  if not remocoes.empty else []

        alt_list = []
        for _, row in alteracoes.iterrows():
            item = {
                "tipo":     str(row.get("TIPO", "")),
                "matricula": str(row.get("MATRÍCULA", "")),
                "nome":     str(row.get("NOME", "")),
            }
            for campo in CAMPOS_MONITORADOS:
                item[f"{campo.lower()}_antes"]  = str(row.get(f"{campo}_ANTES", ""))
                item[f"{campo.lower()}_depois"] = str(row.get(f"{campo}_DEPOIS", ""))
                item[f"{campo.lower()}_mudou"]  = bool(row.get(f"{campo}_MUDOU", False))
            alt_list.append(item)
        resultado["alteracoes"] = alt_list

        # Alerta se mais de 50% removidos (possível problema de chave)
        if len(df_antigo) > 0 and len(remocoes) / len(df_antigo) > 0.5:
            resultado["alerta_chave"] = (
                f"Mais de 50% dos colaboradores marcados como removidos "
                f"({len(remocoes)}/{len(df_antigo)}). "
                "Possível problema de formato na planilha."
            )

    # Salva resultado em JSON
    json_path = pasta_relatorios / f"resultado_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    # Salva snapshot para próxima comparação
    snapshot_path = pasta_snapshots / f"snapshot_{timestamp}_{data_str}.xlsx"
    shutil.copy2(arquivo_novo, snapshot_path)

    return resultado


# ── Geração de relatório Excel para download ───────────────────────────────

def gerar_excel(resultado: dict) -> bytes:
    """Gera relatório Excel formatado a partir dos dados JSON. Retorna bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Alterações {resultado['data']}"

    data_str   = resultado["data"]
    arquivo    = resultado["arquivo"]
    adicoes    = resultado.get("adicoes", [])
    remocoes   = resultado.get("remocoes", [])
    alteracoes = resultado.get("alteracoes", [])

    # ── Cabeçalho ──
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"Relatório de Alterações — {data_str}"
    c.font, c.fill, c.alignment = FONT_TITLE, FILL_TITLE, CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"Arquivo: {arquivo}"
    ws["A2"].font  = Font(name="Arial", size=9, italic=True, color="546E7A")

    ws.merge_cells("A3:I3")
    ws["A3"].value = (
        f"Resumo: {resultado['n_adicoes']} admissões  |  "
        f"{resultado['n_remocoes']} desligamentos  |  "
        f"{resultado['n_alteracoes']} alterações  |  "
        f"Total: {resultado['total']} colaboradores"
    )
    ws["A3"].font = FONT_BOLD

    linha = 5

    def _escrever_secao(ws, linha, titulo, fill, dados, colunas_display, colunas_chave):
        ws.merge_cells(
            start_row=linha, start_column=1,
            end_row=linha,   end_column=len(colunas_display),
        )
        c = ws.cell(row=linha, column=1, value=titulo)
        c.font, c.fill, c.alignment = FONT_SECTION, fill, LEFT
        ws.row_dimensions[linha].height = 20
        linha += 1

        if not dados:
            c = ws.cell(row=linha, column=1, value="Nenhuma ocorrência.")
            c.font = Font(name="Arial", size=10, italic=True, color="78909C")
            ws.merge_cells(
                start_row=linha, start_column=1,
                end_row=linha, end_column=len(colunas_display),
            )
            return linha + 2

        for idx, col in enumerate(colunas_display, 1):
            c = ws.cell(row=linha, column=idx, value=col)
            c.font, c.fill, c.alignment, c.border = FONT_HEADER, FILL_HEADER, CENTER, BORDER
        linha += 1

        for i, item in enumerate(dados):
            for j, chave in enumerate(colunas_chave, 1):
                c = ws.cell(row=linha, column=j, value=item.get(chave, ""))
                c.font = FONT_BODY
                c.alignment = CENTER if j <= 2 else LEFT
                c.border = BORDER
                if i % 2 == 1:
                    c.fill = FILL_ZEBRA
            linha += 1
        return linha + 1

    linha = _escrever_secao(
        ws, linha,
        f"➕  ADMISSÕES ({len(adicoes)})", FILL_ADD,
        adicoes,
        ["Tipo", "Matrícula", "Nome", "Cargo", "Departamento", "Gestor", "Admissão"],
        ["tipo", "matricula", "nome", "cargo", "departamento", "gestor", "admissao"],
    )

    linha = _escrever_secao(
        ws, linha,
        f"➖  DESLIGAMENTOS ({len(remocoes)})", FILL_REM,
        remocoes,
        ["Tipo", "Matrícula", "Nome", "Cargo", "Departamento", "Gestor"],
        ["tipo", "matricula", "nome", "cargo", "departamento", "gestor"],
    )

    # Seção de alterações
    ws.merge_cells(
        start_row=linha, start_column=1,
        end_row=linha, end_column=9,
    )
    c = ws.cell(row=linha, column=1, value=f"✏️   ALTERAÇÕES ({len(alteracoes)})")
    c.font, c.fill, c.alignment = FONT_SECTION, FILL_ALT, LEFT
    ws.row_dimensions[linha].height = 20
    linha += 1

    if not alteracoes:
        c = ws.cell(row=linha, column=1, value="Nenhuma ocorrência.")
        c.font = Font(name="Arial", size=10, italic=True, color="78909C")
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=9)
    else:
        cabecalhos = [
            "Tipo", "Matrícula", "Nome",
            "Cargo Antes",   "Cargo Depois",
            "Depto Antes",   "Depto Depois",
            "Gestor Antes",  "Gestor Depois",
        ]
        for idx, h in enumerate(cabecalhos, 1):
            c = ws.cell(row=linha, column=idx, value=h)
            c.font, c.fill, c.alignment, c.border = FONT_HEADER, FILL_HEADER, CENTER, BORDER
        linha += 1

        campos_alt_keys = [
            "tipo", "matricula", "nome",
            "cargo_antes",    "cargo_depois",
            "departamento_antes", "departamento_depois",
            "gestor_antes",   "gestor_depois",
        ]
        mudou_map = {
            3: "cargo_mudou",  4: "cargo_mudou",
            5: "departamento_mudou", 6: "departamento_mudou",
            7: "gestor_mudou", 8: "gestor_mudou",
        }
        for i, item in enumerate(alteracoes):
            for j, chave in enumerate(campos_alt_keys, 1):
                c = ws.cell(row=linha, column=j, value=item.get(chave, ""))
                c.font = FONT_BODY
                c.alignment = CENTER if j <= 2 else LEFT
                c.border = BORDER
                if i % 2 == 1:
                    c.fill = FILL_ZEBRA
                if j in mudou_map and item.get(mudou_map[j]):
                    c.fill = FILL_DIFF
                    c.font = FONT_BOLD
            linha += 1

    larguras = {"A": 7, "B": 11, "C": 30, "D": 28, "E": 28,
                "F": 26, "G": 26, "H": 26, "I": 26}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Organograma — análise semântica de cargos ─────────────────────────────

def _org_normalizar(texto: str) -> str:
    """Minúsculas, sem acentos, sem espaços duplos."""
    if not texto:
        return ""
    t = str(texto).lower().strip()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t)


# Níveis hierárquicos: 0 = topo, maior = mais operacional.
# Ordem importa: mais específico primeiro.
_NIVEL_RULES: list[tuple[int, list[str]]] = [
    (0, ["conselho de administracao", "presidente executivo", "ceo"]),
    (1, ["superintendente"]),
    (2, ["diretor executivo", "diretor geral", "diretor", "diretora", "vice-presidente"]),
    (3, ["gerente"]),
    (4, ["coordenador", "coordenadora"]),
    (5, ["supervisor", "supervisora", "lider", "lider de equipe", "chefe de"]),
    (6, ["analista", "especialista", "engenheiro", "engenheira",
          "contador", "contadora", "tecnologo", "tecnologa", "desenvolvedor", "desenvolvedora"]),
    (7, ["tecnico", "tecnica"]),
    (8, ["assistente", "auxiliar", "ajudante", "operador", "operadora",
          "motorista", "mecanico", "mecanica", "eletricista", "soldador",
          "zelador", "porteiro", "recepcionista", "almoxarife"]),
    (9, ["estagiario", "estagiaria", "aprendiz", "jovem aprendiz"]),
]


def _nivel_cargo(cargo: str) -> int:
    """Retorna o nível hierárquico (0 = mais alto) com base nas palavras-chave do cargo."""
    c = _org_normalizar(cargo)
    for nivel, kws in _NIVEL_RULES:
        if any(kw in c for kw in kws):
            return nivel
    return 7  # default: nível técnico/operacional


# Palavras ignoradas ao extrair a área do cargo
_AREA_STOPWORDS = {
    "de", "da", "do", "das", "dos", "e", "a", "o", "as", "os",
    "em", "no", "na", "nos", "nas", "para", "por", "com", "um", "uma",
    # títulos que definem nível (removidos para obter só a área)
    "superintendente", "gerente", "diretor", "diretora",
    "coordenador", "coordenadora", "supervisor", "supervisora",
    "analista", "especialista", "engenheiro", "engenheira",
    "tecnico", "tecnica", "assistente", "auxiliar",
    "operador", "operadora", "ajudante", "lider", "chefe",
    "estagiario", "estagiaria", "aprendiz", "contador", "contadora",
    "desenvolvedor", "desenvolvedora", "mecanico", "eletricista",
}


def _area_cargo(cargo: str) -> set[str]:
    """Extrai palavras de área do cargo (remove título de nível e stopwords)."""
    palavras = set(_org_normalizar(cargo).split())
    return palavras - _AREA_STOPWORDS


def _palavras_depto(depto: str) -> set[str]:
    """Extrai palavras relevantes do departamento."""
    stopwords = {"de", "da", "do", "das", "dos", "e", "a", "o", "e"}
    return {w for w in _org_normalizar(depto).split() if w not in stopwords and len(w) > 1}


def _score_compatibilidade(pai: dict, filho: dict) -> int:
    """
    Pontua o quão compatível é um candidato a pai para este filho.
    Leva em conta:
      - Mesmo departamento exato (+10)
      - Palavras do departamento em comum (+4 por palavra)
      - Palavras da área do cargo em comum (+5 por palavra)
      - Área do pai contida na área do filho ou vice-versa (+8)
    """
    score = 0

    dp = _org_normalizar(pai.get("departamento") or "")
    df = _org_normalizar(filho.get("departamento") or "")
    if dp and df:
        if dp == df:
            score += 10
        else:
            score += len(_palavras_depto(dp) & _palavras_depto(df)) * 4

    ap = _area_cargo(pai.get("cargo") or "")
    af = _area_cargo(filho.get("cargo") or "")
    if ap and af:
        score += len(ap & af) * 5
        # Containment: área do pai está toda dentro da área do filho ou vice-versa
        ap_str = " ".join(sorted(ap))
        af_str = " ".join(sorted(af))
        if ap_str and af_str and (ap_str in af_str or af_str in ap_str):
            score += 8

    return score


# ── Estrutura persistida do organograma ───────────────────────────────────────

_ORG_FILE = "organograma_estrutura.json"


def carregar_estrutura_org(data_dir: Path) -> dict:
    """
    Carrega posições manuais e relações adicionais salvas.
    Retorna:
        {
          "posicoes": {chave: parent_chave},
          "relacoes": [{"filho": chave, "pai": chave}, ...]
        }
    """
    try:
        path = data_dir / _ORG_FILE
        if path.exists():
            with open(path, encoding="utf-8") as f:
                dados = json.load(f)
            return {
                "posicoes": dados.get("posicoes", {}),
                "relacoes": dados.get("relacoes_adicionais", []),
            }
    except Exception:
        pass
    return {"posicoes": {}, "relacoes": []}


def _gravar_org_json(
    data_dir: Path, posicoes_novas: dict, usuario: str,
    relacoes_novas: list | None = None,
    acao: str = "",
) -> None:
    """Merge incremental com histórico de snapshots (últimos 30)."""
    path   = data_dir / _ORG_FILE
    dados: dict = {"versao": 3, "posicoes": {}, "relacoes_adicionais": [], "snapshots": []}
    if path.exists():
        try:
            with open(path, encoding="utf-8") as f:
                dados = json.load(f)
        except Exception:
            pass
    dados.setdefault("posicoes", {})
    dados.setdefault("relacoes_adicionais", [])
    dados.setdefault("snapshots", [])

    # ── Snapshot do estado ANTES da alteração (permite desfazer) ─────────
    dados["snapshots"].append({
        "ts":       datetime.now().isoformat(),
        "usuario":  usuario,
        "acao":     acao or "Alteração manual",
        "posicoes": dict(dados["posicoes"]),
        "relacoes": list(dados["relacoes_adicionais"]),
    })
    if len(dados["snapshots"]) > 30:
        dados["snapshots"] = dados["snapshots"][-30:]

    # ── Aplica as mudanças ────────────────────────────────────────────────
    dados["posicoes"].update(posicoes_novas)
    if relacoes_novas is not None:
        filhos_novos = {r["filho"] for r in relacoes_novas}
        dados["relacoes_adicionais"] = [
            r for r in dados["relacoes_adicionais"]
            if r.get("filho") not in filhos_novos
        ]
        dados["relacoes_adicionais"].extend(relacoes_novas)

    dados["versao"]         = 3
    dados["atualizado_em"]  = datetime.now().isoformat()
    dados["atualizado_por"] = usuario
    data_dir.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def salvar_posicao_org(
    data_dir: Path, chave: str, parent_chave: str,
    usuario: str = "", acao: str = ""
) -> None:
    """Salva (ou atualiza) a posição hierárquica de UM colaborador."""
    _gravar_org_json(data_dir, {chave: parent_chave}, usuario, acao=acao)


def salvar_posicoes_org(
    data_dir: Path, posicoes: dict, usuario: str = ""
) -> None:
    """Salva/atualiza posições de MÚLTIPLOS colaboradores de uma vez."""
    _gravar_org_json(data_dir, posicoes, usuario, acao="Aceitou sugestões automáticas em lote")


def salvar_relacoes_org(
    data_dir: Path, filho: str, pais_adicionais: list,
    usuario: str = "", acao: str = ""
) -> None:
    """Salva/substitui os gestores adicionais (múltiplos) de UM colaborador."""
    relacoes = [
        {"filho": filho, "pai": pai}
        for pai in pais_adicionais
        if pai and pai != filho
    ]
    _gravar_org_json(data_dir, {}, usuario, relacoes, acao=acao)


# ── Validação e histórico ─────────────────────────────────────────────────────

def validar_posicao_org(posicoes: dict, chave: str, novo_pai: str) -> tuple[bool, str]:
    """
    Verifica se definir chave → novo_pai criaria um ciclo na hierarquia.
    Retorna (ok: bool, mensagem_de_erro: str).
    """
    if not novo_pai or novo_pai in ("", "__ROOT__"):
        return True, ""
    if novo_pai == chave:
        return False, "Um nó não pode reportar a si mesmo."
    visitados: set[str] = set()
    atual = novo_pai
    while atual and atual not in ("", "__ROOT__"):
        if atual == chave:
            return False, (
                "A operação criaria um ciclo: o superior selecionado "
                "já está subordinado a esta pessoa na hierarquia atual."
            )
        if atual in visitados:
            break  # já há ciclo pré-existente; para para não travar
        visitados.add(atual)
        atual = posicoes.get(atual, "")
    return True, ""


def carregar_historico_org(data_dir: Path) -> list:
    """
    Retorna os snapshots salvos (metadados apenas, sem o conteúdo completo).
    Ordem: mais recente primeiro.
    """
    try:
        path = data_dir / _ORG_FILE
        if path.exists():
            with open(path, encoding="utf-8") as f:
                dados = json.load(f)
            snaps = dados.get("snapshots", [])
            return [
                {
                    "ts":         s["ts"],
                    "usuario":    s.get("usuario", ""),
                    "acao":       s.get("acao", ""),
                    "n_posicoes": len(s.get("posicoes", {})),
                }
                for s in reversed(snaps)
            ]
    except Exception:
        pass
    return []


def restaurar_snapshot_org(data_dir: Path, ts: str, usuario: str) -> tuple[bool, str]:
    """
    Restaura o estado de um snapshot específico (identificado pelo timestamp).
    Salva o estado atual como um snapshot de backup antes de restaurar.
    Retorna (ok: bool, mensagem: str).
    """
    try:
        path = data_dir / _ORG_FILE
        if not path.exists():
            return False, "Arquivo de estrutura não encontrado."
        with open(path, encoding="utf-8") as f:
            dados = json.load(f)
        snaps = dados.get("snapshots", [])
        snap  = next((s for s in snaps if s["ts"] == ts), None)
        if not snap:
            return False, f"Snapshot '{ts}' não encontrado."

        # Salva o estado atual como backup antes de restaurar
        dados.setdefault("snapshots", []).append({
            "ts":       datetime.now().isoformat(),
            "usuario":  usuario,
            "acao":     f"(backup antes de restaurar ponto {ts[:16]})",
            "posicoes": dict(dados.get("posicoes", {})),
            "relacoes": list(dados.get("relacoes_adicionais", [])),
        })

        # Restaura
        dados["posicoes"]            = snap["posicoes"]
        dados["relacoes_adicionais"] = snap.get("relacoes", [])
        dados["versao"]              = 3
        dados["atualizado_em"]       = datetime.now().isoformat()
        dados["atualizado_por"]      = usuario

        if len(dados["snapshots"]) > 30:
            dados["snapshots"] = dados["snapshots"][-30:]

        with open(path, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
        return True, "Restaurado com sucesso."
    except Exception as exc:
        return False, str(exc)


def importar_estrutura_org(
    data_dir: Path, posicoes: dict, relacoes: list, usuario: str
) -> None:
    """
    Importa/substitui a estrutura completa do organograma.
    Preserva o histórico de snapshots existente.
    """
    path  = data_dir / _ORG_FILE
    dados: dict = {"versao": 3, "posicoes": {}, "relacoes_adicionais": [], "snapshots": []}
    if path.exists():
        try:
            with open(path, encoding="utf-8") as f:
                dados = json.load(f)
        except Exception:
            pass
    dados.setdefault("snapshots", [])

    # Backup antes de importar
    dados["snapshots"].append({
        "ts":       datetime.now().isoformat(),
        "usuario":  usuario,
        "acao":     "Backup antes de importação manual",
        "posicoes": dict(dados.get("posicoes", {})),
        "relacoes": list(dados.get("relacoes_adicionais", [])),
    })
    if len(dados["snapshots"]) > 30:
        dados["snapshots"] = dados["snapshots"][-30:]

    dados["posicoes"]            = posicoes
    dados["relacoes_adicionais"] = relacoes
    dados["versao"]              = 3
    dados["atualizado_em"]       = datetime.now().isoformat()
    dados["atualizado_por"]      = usuario
    data_dir.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def gerar_organograma_hibrido(
    colaboradores: list[dict],
    posicoes: dict,
    relacoes: list | None = None,
) -> dict:
    """
    Constrói a lista de nós do organograma mesclando:
      - Posições manuais salvas (posicoes: {chave → parent_chave})
      - Análise semântica de cargos para colaboradores ainda sem posição

    Cada colaborador deve ter o campo "chave" (chave do snapshot).
    A raiz virtual "__ROOT__" (ALISEO SA) é sempre inserida no topo.

    Retorna:
        {
          "nodes":      list[dict],   # formato para d3.stratify
          "n_manual":   int,          # nós com posição confirmada pelo usuário
          "n_sugerido": int,          # nós com posição sugerida automaticamente
        }
    """
    if not colaboradores:
        return {"nodes": [], "n_manual": 0, "n_sugerido": 0}

    por_chave       = {c["chave"]: c for c in colaboradores}
    chaves_com_pos  = set(posicoes) & set(por_chave)
    chaves_sem_pos  = set(por_chave) - chaves_com_pos

    # Nó raiz virtual — sempre presente, nunca editável
    nodes: list[dict] = [{
        "id": "__ROOT__", "parentId": "",
        "name": "ALISEO SA", "cargo": "Conselho de Administração",
        "departamento": "", "tipo": "", "nivel": -1,
        "manual": True, "chave": "__ROOT__",
    }]

    # ── 1. Nós com posição salva manualmente ──────────────────────────────
    for chave in chaves_com_pos:
        c      = por_chave[chave]
        parent = posicoes[chave]
        # Pai foi desligado ou é vazio → fica direto abaixo do Conselho
        if parent and parent != "__ROOT__" and parent not in por_chave:
            parent = "__ROOT__"
        if not parent:
            parent = "__ROOT__"
        nivel = _nivel_cargo(c.get("cargo", ""))
        nodes.append({
            "id":           chave,
            "parentId":     parent,
            "name":         c.get("nome", ""),
            "cargo":        c.get("cargo", ""),
            "departamento": c.get("departamento", ""),
            "tipo":         c.get("tipo", ""),
            "nivel":        nivel,
            "manual":       True,
            "chave":        chave,
        })

    # ── 2. Nós sem posição: algoritmo semântico ───────────────────────────
    # Pool de candidatos a pai: nós já posicionados (exceto a raiz virtual)
    candidatos: list[dict] = [
        {"id": n["id"], "cargo": n["cargo"],
         "departamento": n["departamento"], "nivel": n["nivel"]}
        for n in nodes if n["id"] != "__ROOT__"
    ]

    sem_pos = sorted(
        [{**por_chave[ch], "nivel": _nivel_cargo(por_chave[ch].get("cargo", ""))}
         for ch in chaves_sem_pos],
        key=lambda x: (x["nivel"], _org_normalizar(x.get("nome", ""))),
    )

    for c in sem_pos:
        nivel    = c["nivel"]
        superiores = [x for x in candidatos if x["nivel"] < nivel]
        if superiores:
            nivel_alvo = max(x["nivel"] for x in superiores)
            cands  = [x for x in superiores if x["nivel"] == nivel_alvo]
            melhor = max(cands, key=lambda x: _score_compatibilidade(x, c))
            parent_chave = melhor["id"]
        else:
            parent_chave = "__ROOT__"    # sem superior → direto no Conselho

        nodes.append({
            "id":           c["chave"],
            "parentId":     parent_chave,
            "name":         c.get("nome", ""),
            "cargo":        c.get("cargo", ""),
            "departamento": c.get("departamento", ""),
            "tipo":         c.get("tipo", ""),
            "nivel":        nivel,
            "manual":       False,       # sugestão automática, não confirmada
            "chave":        c["chave"],
        })
        # Entra no pool para os próximos colaboradores da mesma rodada
        candidatos.append({
            "id": c["chave"], "cargo": c.get("cargo", ""),
            "departamento": c.get("departamento", ""), "nivel": nivel,
        })

    n_manual   = sum(1 for n in nodes if n.get("manual")  and n["id"] != "__ROOT__")
    n_sugerido = sum(1 for n in nodes if not n.get("manual"))

    # ── 3. Arestas extras (múltiplos gestores / linhas pontilhadas) ───────────
    chaves_validas = {n["id"] for n in nodes}
    arestas_extras = [
        {"filho": r["filho"], "pai": r["pai"]}
        for r in (relacoes or [])
        if r.get("filho") in chaves_validas
        and r.get("pai") in chaves_validas
        and r.get("filho") != r.get("pai")
    ]

    return {
        "nodes":          nodes,
        "n_manual":       n_manual,
        "n_sugerido":     n_sugerido,
        "arestas_extras": arestas_extras,
    }
